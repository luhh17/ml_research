import os
import pickle as pkl
import pandas as pd
import numpy as np
from exp.post_exp import Post_Exp
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from utils.post_analysis import convert_array, backtest, balance_weight, calculate_turnover, calculate_sr, calculate_port_ret, set_number_format
from prepare_dataset.myDataset import split_signal_return_mask

import pdb

config_var = ['model', 'window_length', 'pred_length', 'dec_len_from_input', 'embed_dim', 'hidden_dim',
                                              'num_enc_layers', 'num_dec_layers', 'num_heads']
adv_limit = 0.01
adv_limit_hold = 0.05
std_cutoff = 1
capital_tot = 1e8
trading_cost_rate = 0.0008
period_start='2018-01-01'
period_end='2023-07-01'

# 计算ensemble模型后的结果
def ensemble_result(path_name='ensemble_model_file', num_model=-1, strategy='bagging', date_var='date', ret_var='ret_open', id_var='stkcd',universe='top1000'):
    
    path = f'../{path_name}'
    dict_name = f'{path_name}_{universe}_res_dict.pkl'
    output_name = f'{path_name}_{universe}_result.xlsx'
    port_ret_name= f'{path_name}_{universe}_result_longport_ret.xlsx'
    model_list = np.sort(os.listdir(path))

    model_name_path_dict = {model_name: os.path.join(path, model_name) for model_name in model_list}

    long_port_dfs=[]
    result_dfs = []
    mean_by_year = []
    sr_by_year = []
    dd_by_year=[]
    to_by_year = []
    output_model_list = []

    if not os.path.exists(f'../report_{path_name}'):
        os.makedirs(f'../report_{path_name}')

    if os.path.exists(f'../report_{path_name}/{dict_name}'):
        res_dict = pkl.load(open(f'../report_{path_name}/{dict_name}', 'rb'))
    else:
        res_dict = {}

    for path in model_name_path_dict.values():
        if path.split('/')[-1] in list(res_dict.keys()):
            continue

        model_idx = np.sort(os.listdir(path))
        model_idx = [m for m in model_idx if os.path.isdir(os.path.join(path, m))]
        pred_list = []
        
        if len(model_idx) == 0:
            model_idx = [-1]
        for idx in model_idx:
            if int(idx) >= num_model and num_model != -1:
                break
            idx_path = path
            if model_idx[0] != -1:
                idx_path += f'/{idx}'
            file_list = os.listdir(idx_path)
            pred_file_list = [f for f in file_list if f.endswith('_pred_result')]
            if '2023_pred_result' not in pred_file_list:
                continue
            pred_df = []
            for pred in pred_file_list:
                df = pd.read_pickle(f'{idx_path}/{pred}')
                df = df.rename(columns={'weight': 'pred'})
                pred_df.append(df)
            pred_df = pd.concat(pred_df, axis=0)
            if 'ret' in pred_df.columns:
                pred_df = pred_df.drop(columns=['ret'])
            pred_df[date_var] = pd.to_datetime(pred_df[date_var], format='%Y%m%d')
            pred_df = pred_df[pred_df['mask'] == True]
            pred_df = pred_df.sort_values(by=[date_var, id_var])
            pred = pred_df['pred'].values
            pred_list.append(pred)


        if len(pred_list) == 0:
            continue
        output_model_list.append(path.split('/')[-1])
        pred_list = np.array(pred_list)
        print(path.split('/')[-1])
        pred_list = np.mean(pred_list, axis=0)
        pred_df['pred'] = pred_list
        pred_df=pred_df[(pred_df['date']>=period_start) & (pred_df['date']<=period_end)]
        
        trading_adv_df = pd.read_pickle('/mnt/HDD16TB/backtest_data/trading_adv_info.pkl')
        
        if universe=='top1000':
            trading_adv_df['year'] = trading_adv_df['date'].dt.year
            trading_adv_df_top1000 = pd.DataFrame()
            for year in trading_adv_df['year'].unique():
                # Filter data for the current year
                year_data = trading_adv_df[trading_adv_df['year'] == year]
                first_date_of_year = year_data['date'].min()
                year_data_first_date = year_data[year_data['date'] == first_date_of_year].copy()
                year_data_first_date.loc[:,'mktcap_r'] = year_data_first_date['size'].rank(ascending=False, method='min')
                top1000_universe=year_data_first_date[year_data_first_date['mktcap_r'] <= 1000]['stkcd']
                trading_adv_df_top1000 = pd.concat([trading_adv_df_top1000, year_data[year_data['stkcd'].isin(top1000_universe)]])
                del year_data,year_data_first_date

            trading_adv_df=trading_adv_df_top1000.copy()
            # trading_adv_df['mktcap_r'] = trading_adv_df.groupby('date')['size'].rank(ascending=False, method='min')
            # trading_adv_df = trading_adv_df[trading_adv_df['mktcap_r'] <= 1000].reset_index(drop=True)

        # return_df = pd.read_pickle('/mnt/HDD16TB/backtest_data/Ashare_Stock_return.pkl')
        return_df=pd.read_pickle('/mnt/HDD16TB/backtest_data/ret_decomposed.pkl')
        market_df = pd.read_pickle('/mnt/HDD16TB/backtest_data/CSI_500.pkl')
        
        index_df = pd.read_pickle('/mnt/HDD16TB/backtest_data/IDX_Smprat.pkl')[['Indexcd','Enddt','Stkcd','Weight']]
        index_df.rename(columns={'Enddt':'date','Stkcd':'stkcd'},inplace=True)

        return_df = pd.merge(return_df, index_df, on=[date_var, id_var], how='left')
        return_df = pd.merge(return_df, market_df, on=date_var, how='left')
        return_df['exret'] = return_df['ret_close'] - return_df['mktret']
        ret_adv_df = pd.merge(return_df, trading_adv_df, on=[date_var, id_var], how='inner')
        market_df.set_index(date_var, inplace=True)
        pred_df.rename(columns={'pred': 'Raw'}, inplace=True)
        if pred_df[id_var].dtype != 'str':
            pred_df[id_var] = pred_df[id_var].astype(str).str.zfill(6)
        
        
        # Combine and pre-process data
        predict_df = pd.merge(pred_df, ret_adv_df, on=[date_var, id_var], how='inner')
        date_series = predict_df[date_var].unique()
        predict_df.set_index([date_var, id_var], inplace=True)
        mkt_df_subset = market_df.loc[date_series]

        array, mask, _, columns = convert_array(predict_df)
        signal = array[..., columns.get_loc('Raw')].astype(float)
        ret = array[..., columns.get_loc(ret_var)].astype(float)
        stock_adv = array[..., columns.get_loc('adv')].astype(float)
        mask_buy = array[..., columns.get_loc('mask_buy')].astype(bool) * mask
        mask_sell = array[..., columns.get_loc('mask_sell')].astype(bool) * mask
        market = mkt_df_subset['mktret'].to_numpy()
        stock_adv_limit = adv_limit * stock_adv
        stock_adv_limit_hold = adv_limit_hold * stock_adv

        ret_rb15_bc, turnover_rb15_bc, _ = backtest(signal, ret, stock_adv_limit, stock_adv_limit_hold, mask_buy,
                                                    mask_sell, capital_tot,
                                                    shorting=True, market=market, rebalance=1.5, cost_rate=0,
                                                    std_cutoff=std_cutoff)
        ret_rb15_ac, turnover_rb15_ac, _ = backtest(signal, ret, stock_adv_limit, stock_adv_limit_hold, mask_buy,
                                                    mask_sell, capital_tot,
                                                    shorting=True, market=market, rebalance=1.5,
                                                    cost_rate=trading_cost_rate, std_cutoff=std_cutoff)
        ret_rb1_bc, turnover_rb1_bc, _ = backtest(signal, ret, stock_adv_limit, stock_adv_limit_hold, mask_buy,
                                                mask_sell, capital_tot,
                                                shorting=True, market=market, rebalance=1, cost_rate=0,
                                                std_cutoff=std_cutoff)
        ret_rb1_ac, turnover_rb1_ac, _ = backtest(signal, ret, stock_adv_limit, stock_adv_limit_hold, mask_buy,
                                                mask_sell, capital_tot,
                                                shorting=True, market=market, rebalance=1,
                                                cost_rate=trading_cost_rate, std_cutoff=std_cutoff)
        ret_long_bc, turnover_long_bc, _ = backtest(signal, ret, stock_adv_limit, stock_adv_limit_hold, mask_buy,
                                                    mask_sell, capital_tot,
                                                    cost_rate=0, std_cutoff=std_cutoff)
        ret_long_ac, turnover_long_ac, _ = backtest(signal, ret, stock_adv_limit, stock_adv_limit_hold, mask_buy,
                                                    mask_sell, capital_tot,
                                                    cost_rate=trading_cost_rate, std_cutoff=std_cutoff)

        # Balance weights
        predict_df['Balanced'] = balance_weight(predict_df, 'Raw')
        predict_df['Long'] = predict_df['Balanced'] * (predict_df['Balanced'] > 0)
        predict_df['Short'] = predict_df['Balanced'] * (predict_df['Balanced'] < 0)
        weight_300=predict_df[predict_df['Indexcd']=='000300'].groupby(level=0)['Long'].sum()
        weight_500=predict_df[predict_df['Indexcd']=='000905'].groupby(level=0)['Long'].sum()
        weight_others=1-weight_300-weight_500
        
        # Calculate size and adv average ranks
        rank_df = pd.DataFrame(index=predict_df.index)
        rank_df[['Raw', 'Balanced', 'Long', 'Short']] = abs(predict_df[['Raw', 'Balanced', 'Long', 'Short']])
        rank_df['Raw'] /= rank_df['Raw'].groupby(date_var).sum()
        rank_df['Balanced'] /= 2
        rank_df['size'] = predict_df['size'].groupby(date_var).rank(ascending=False)
        rank_df['adv'] = predict_df['adv'].groupby(date_var).rank(ascending=False)
        size_avg_rank = (
                rank_df[['Raw', 'Balanced', 'Long', 'Short']] * rank_df['size'].to_numpy().reshape(-1, 1)).groupby(
            date_var).sum().mean()
        adv_avg_rank = (
                rank_df[['Raw', 'Balanced', 'Long', 'Short']] * rank_df['adv'].to_numpy().reshape(-1, 1)).groupby(
            date_var).sum().mean()

        # Calculate statistics
        turnover_df = calculate_turnover(predict_df[['Raw', 'Balanced', 'Long', 'Short']])
        turnover_df['ADV_rb1.5_bc'] = turnover_rb15_bc
        turnover_df['ADV_rb1.5_ac'] = turnover_rb15_ac
        turnover_df['ADV_rb1_bc'] = turnover_rb1_bc
        turnover_df['ADV_rb1_ac'] = turnover_rb1_ac
        turnover_df['ADV_long_bc'] = turnover_long_bc
        turnover_df['ADV_long_ac'] = turnover_long_ac

        port_absret_df,port_ret_df = calculate_port_ret(predict_df[['ret_close_to_open','ret_open_to_close', 'mktret','Raw', 'Balanced', 'Long', 'Short']])
        
        
        port_ret_df['ADV_rb1.5_bc'] = ret_rb15_bc
        port_ret_df['ADV_rb1.5_ac'] = ret_rb15_ac
        port_ret_df['ADV_rb1_bc'] = ret_rb1_bc
        port_ret_df['ADV_rb1_ac'] = ret_rb1_ac
        port_ret_df['ADV_long_bc'] = ret_long_bc - market
        port_ret_df['ADV_long_ac'] = ret_long_ac - market
        

        index_ret=pd.read_csv('/mnt/HDD16TB/backtest_data/indexret.csv',index_col='date')
        index_ret.index=pd.to_datetime(index_ret.index)
        index_ret=index_ret.loc[port_ret_df.index,:]
        mkt_ret=market_df.loc[port_ret_df.index,:]
        fig, axes = plt.subplots(1, 2, figsize=(18, 9), dpi=300)
        ret = port_ret_df[['Long','ac_Long']].cumsum()
        ret.plot(ax=axes[0],label='Cumulative exmkt Returns')   
        index_ret.cumsum().plot(ax=axes[0])  
        axes[0].set_title('Cumulative exmkt Returns') 
        axes[0].legend()
        ret = port_absret_df[['Long','ac_Long']].cumsum()
        ret.plot(ax=axes[1],label='Cumulative absolute Returns')   
        index_ret.cumsum().plot(ax=axes[1])  
        axes[1].set_title('Cumulative absolute Returns')
        axes[1].legend()
        model_name=path.split('/')[-1]
        fig.savefig(f'../report_{path_name}/{model_name}_{universe}.png') 


        long_port_df=pd.concat([port_absret_df['Long'],mkt_ret,index_ret,weight_300,weight_500,weight_others],axis=1)
        long_port_df.columns=['bc_long_daily_ret','mkt_ret','300_daily_ret','500_daily_ret','300_weight','500_weight','others_weight']

        result_df = calculate_sr(port_ret_df)
        result_df.loc['TO'] = turnover_df.mean()
        result_df.loc['Size_rank', ['Raw', 'Balanced', 'Long', 'Short']] = size_avg_rank
        result_df.loc['ADV_rank', ['Raw', 'Balanced', 'Long', 'Short']] = adv_avg_rank
        result_by_year_df = port_ret_df.groupby(port_ret_df.index.year).apply(calculate_sr)
        to_by_year_df = turnover_df.groupby(turnover_df.index.year).mean()

        # Add results
        long_port_dfs.append(long_port_df)
        result_dfs.append(result_df)
        mean_by_year.append(result_by_year_df.loc[:, 'mean', :])
        sr_by_year.append(result_by_year_df.loc[:, 'SR', :])
        dd_by_year.append(result_by_year_df.loc[:, 'max_dd', :])
        to_by_year.append(to_by_year_df)

    per_model_df = [[result_dfs[i], mean_by_year[i], sr_by_year[i],dd_by_year[i], to_by_year[i]] for i in range(len(output_model_list))]
    if len(output_model_list) != 0:
        new_res_dict = dict(zip(output_model_list, per_model_df))
        res_dict.update(new_res_dict)
    pkl.dump(res_dict, open(f'../report_{path_name}/{dict_name}', 'wb'))
    result_dfs = [v[0] for v in list(res_dict.values())]
    mean_by_year = [v[1] for v in list(res_dict.values())]
    sr_by_year = [v[2] for v in list(res_dict.values())]
    dd_by_year=[v[3] for v in list(res_dict.values())]
    to_by_year = [v[4] for v in list(res_dict.values())]
    output_model_list = list(res_dict.keys())
    # Combine results
    result = pd.concat(result_dfs, keys=output_model_list, names=['Model', 'Stat'])
    mean_by_year = pd.concat(mean_by_year, keys=output_model_list, names=['Model', 'year'])
    sr_by_year = pd.concat(sr_by_year, keys=output_model_list, names=['Model', 'year'])
    dd_by_year = pd.concat(dd_by_year, keys=output_model_list, names=['Model', 'year'])
    to_by_year = pd.concat(to_by_year, keys=output_model_list, names=['Model', 'year'])

    # Save results
    
    port_result_save_file=os.path.join(f'../report_{path_name}/{port_ret_name}')
    with pd.ExcelWriter(port_result_save_file) as writer1:
        for i,long_port_df in enumerate(long_port_dfs):
            long_port_df.to_excel(writer1, sheet_name=output_model_list[i])
            set_number_format(writer1.sheets[output_model_list[i]], ((1, 1), (long_port_df.shape[0]+1, 2)), 'yyyy-mm-dd')
            set_number_format(writer1.sheets[output_model_list[i]], ((1, 2), (long_port_df.shape[0]+1, 9)), '0.00%')


    result_save_file = os.path.join(f'../report_{path_name}/{output_name}')
    with pd.ExcelWriter(result_save_file) as writer:
    # with load_workbook(filename=result_save_file) as writer:
    # writer = load_workbook(filename=result_save_file)
        result.to_excel(writer, sheet_name='result')
        mean_by_year.to_excel(writer, sheet_name='Mean by year')
        sr_by_year.to_excel(writer, sheet_name='SR by year')
        dd_by_year.to_excel(writer, sheet_name='Drawdown by year')
        to_by_year.to_excel(writer, sheet_name='TO by year')

        # Set number format
        n_models = len(model_list)
        n_rows = mean_by_year.shape[0]
        n_columns = result.shape[1]
        for i in range(n_models):
            set_number_format(writer.sheets['result'], ((9 * i + 2, 3), (9 * i + 4, 3 + n_columns)), '0.00%')
            set_number_format(writer.sheets['result'], ((9 * i + 4, 3), (9 * i + 5, 3 + n_columns)), '0.00')
            set_number_format(writer.sheets['result'], ((9 * i + 5, 3), (9 * i + 6, 3 + n_columns)), '0.00%')
            set_number_format(writer.sheets['result'], ((9 * i + 6, 3), (9 * i + 11, 3 + n_columns)), '0.00')
        set_number_format(writer.sheets['Mean by year'], ((2, 3), (2 + n_rows, 3 + n_columns)), '0.00%')
        set_number_format(writer.sheets['SR by year'], ((2, 3), (2 + n_rows, 3 + n_columns)), '0.00')
        set_number_format(writer.sheets['Drawdown by year'], ((2, 3), (2 + n_rows, 3 + n_columns)), '0.00%')
        set_number_format(writer.sheets['TO by year'], ((2, 3), (2 + n_rows, 3 + n_columns)), '0.00')

        # Set auto column width
        writer.sheets['result'].column_dimensions['A'].auto_size = True
        writer.sheets['Mean by year'].column_dimensions['A'].auto_size = True
        writer.sheets['SR by year'].column_dimensions['A'].auto_size = True
        writer.sheets['Drawdown by year'].column_dimensions['A'].auto_size = True
        writer.sheets['TO by year'].column_dimensions['A'].auto_size = True


if __name__ == '__main__':
    file_list=['model_selected','ensemble_model_file_basic','ensemble_model_file_intraday_only_t12','ensemble_model_file_intraday_only_t15_os']
    file_path=file_list[1]

    ensemble_result(file_path,universe='top1000')
